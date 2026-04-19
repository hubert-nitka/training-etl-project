"""
Generate workout Excel file with exercises, weights from history, and checkboxes for each set
"""
import os
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import json
from datetime import date
from src.utils import connect_to_database, log
from config import WORKOUT_XLSX_PATH


def get_last_weight_used(engine, exercise_id, reps):
    """
    Get the last weight used for an exercise with given reps
    Returns the most recent weight from workout history
    """
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                SELECT se.weight_used, ws.session_date
                FROM session_exercises se
                JOIN workout_sessions ws ON se.session_id = ws.session_id
                WHERE se.exercise_id = :exercise_id
                  AND se.reps_completed = :reps
                ORDER BY ws.session_date DESC
                LIMIT 1
            """),
            {"exercise_id": exercise_id, "reps": reps}
        )
        
        row = result.fetchone()
        return row[0] if row else None
    
def get_weight_for_reps(engine, exercise_id, reps):
    if isinstance(reps, int):
        return get_last_weight_used(engine, exercise_id, reps)
    else:
        # AMRAP
        return get_last_weight_used(engine, exercise_id, -1)


def generate_workout_excel(plan_name, day_of_week, output_file="workout.xlsx"):
    """
    Generate Excel workout tracker for a specific day
    
    Args:
        plan_name: Name of the training plan (e.g., "Plan Luty 2025")
        day_of_week: Day to generate (e.g., "Monday", "Wednesday", "Friday")
        output_file: Output Excel filename
    """
    
    engine = connect_to_database()
    
    try:
        with engine.connect() as conn:
            # Get plan_id
            result = conn.execute(
                text("SELECT plan_id FROM training_plans WHERE plan_name = :plan_name"),
                {"plan_name": plan_name}
            )
            
            row = result.fetchone()
            if not row:
                log(f"Plan '{plan_name}' not found!", level="ERROR", echo=True)
                return
            
            plan_id = row[0]
            
            # Get exercises for this day
            result = conn.execute(
                text("""
                    SELECT 
                        e.exercise_id,
                        e.exercise_name,
                        pe.warmup_sets,
                        pe.working_sets,
                        pe.reps,
                        pe.rest_between_sets_min,
                        pe.rest_between_sets_max,
                        pe.rest_after_exercise_min,
                        pe.rest_after_exercise_max,
                        pe.trainer_note
                    FROM plan_exercises pe
                    JOIN exercises e ON pe.exercise_id = e.exercise_id
                    WHERE pe.plan_id = :plan_id AND pe.day_of_week = :day_of_week
                    ORDER BY pe.plan_exercise_id
                """),
                {"plan_id": plan_id, "day_of_week": day_of_week}
            )
            
            exercises = result.fetchall()
            
            if not exercises:
                log(f"No exercises found for {day_of_week} in '{plan_name}'", level="ERROR", echo=True)
                return
        
        # Create workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = f"{day_of_week}"
        checkbox_dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
        sheet.add_data_validation(checkbox_dv)
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        
        
        # Check for max weight columns needed

        max_weight_cols = max(
            len(json.loads(ex[4])) if ex[4] else 1
            for ex in exercises
        )

        # Add checkboxes columns (one per set)
        max_total_sets = max(
            (ex[2] or 0) + ex[3] for ex in exercises
        )

        # Header row
        headers = [
            "Exercise"
            ] + [f"W{i+1} (kg)" for i in range(max_weight_cols)] + [
            "Warmup sets", "Working sets", "Reps", "Rest between reps", 
            "Rest after exercise", "Trainer note", "Notes"
        ]

        checkbox_start_col = len(headers) + 1
        for i in range(max_total_sets):
            col_letter = get_column_letter(checkbox_start_col + i)
            sheet.column_dimensions[col_letter].width = 4

        # Set column widths
        for i in range(len(headers)):
            col_letter = get_column_letter(i + 1)
            
            if headers[i] == "Exercise":
                sheet.column_dimensions[col_letter].width = 50
            elif "W" in headers[i]:
                sheet.column_dimensions[col_letter].width = 10
            else:
                sheet.column_dimensions[col_letter].width = 18
    
        col_map = {name: idx+1 for idx, name in enumerate(headers)}
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Add checkbox headers
        for i in range(max_total_sets):
            cell = sheet.cell(row=1, column=checkbox_start_col + i)
            cell.value = "✓"
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Add exercises
        current_row = 2
        
        for exercise in exercises:
            (exercise_id, exercise_name, warmup_sets, working_sets, 
             reps_json, rest_min, rest_max, 
             rest_after_min, rest_after_max, trainer_note) = exercise
            
            try:
                reps_list = json.loads(reps_json) if reps_json else []
            except (json.JSONDecodeError, TypeError):
                reps_list = []
            
            # Check for descending series and get weight suggestion from history (use first rep count)
            if len(set(reps_list)) == 1:
                weight_display = [get_weight_for_reps(engine, exercise_id, reps_list[0]) or ""]
            else:
                weight_display = [
                    get_weight_for_reps(engine, exercise_id, reps) or ""
                    for reps in reps_list
                ]
            
            # Format rest times
            rest_between = f"{rest_min}-{rest_max}" if rest_min and rest_max and rest_min != rest_max else str(rest_min) if rest_min else ""
            rest_after = f"{rest_after_min}-{rest_after_max}" if rest_after_min and rest_after_max and rest_after_min != rest_after_max else str(rest_after_min) if rest_after_min else ""
            
            # Add exercise row
            sheet.cell(row=current_row, column=col_map['Exercise']).value = exercise_name
            for i in range(max_weight_cols):
                col_name = f"W{i+1} (kg)"
                value = weight_display[i] if i < len(weight_display) else ""
                sheet.cell(row=current_row, column=col_map[col_name]).value = value
            sheet.cell(row=current_row, column=col_map['Warmup sets']).value = warmup_sets or 0
            sheet.cell(row=current_row, column=col_map['Working sets']).value = working_sets
            sheet.cell(row=current_row, column=col_map['Reps']).value = ", ".join(str(r) for r in reps_list) if reps_list else ""
            sheet.cell(row=current_row, column=col_map['Rest between reps']).value = rest_between
            sheet.cell(row=current_row, column=col_map['Rest after exercise']).value = rest_after
            sheet.cell(row=current_row, column=col_map['Trainer note']).value = trainer_note
            
            # Add borders
            for col in range(1, len(headers) + 1):
                sheet.cell(row=current_row, column=col).border = thin_border
                sheet.cell(row=current_row, column=col).alignment = Alignment(vertical="center")
            
            # Center align numeric columns
            for col_name, col_idx in col_map.items():
                if "W" in col_name or col_name in ["Warmup sets", "Working sets"]:
                    sheet.cell(row=current_row, column=col_idx).alignment = center_align
            
            # Add checkboxes for each set (warmup + working)
            total_sets = (warmup_sets or 0) + working_sets
            for set_num in range(total_sets):
                checkbox_col = checkbox_start_col + set_num
                cell = sheet.cell(row=current_row, column=checkbox_col)
                cell.value = "☐"  # Empty checkbox
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(size=14)
                checkbox_dv.add(cell)
            
            current_row += 1
        
        # Save workbook
        wb.save(str(output_file))
        log(f"Workout Excel generated: {output_file}")
        log(f"Plan: {plan_name}")
        log(f"Day: {day_of_week}")
        log(f"Exercises: {len(exercises)}")
        
    finally:
        engine.dispose()